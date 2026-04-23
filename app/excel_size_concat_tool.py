
import os
import re
import math
import traceback
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from typing import List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter


APP_TITLE = "Excel 款色尺码拼接工具"
HEADER_SCAN_MAX_ROW = 20
PREVIEW_DATA_ROWS = 20
GROUPS_PER_PAGE = 5
COLOR_ALIASES = {"款色", "款号"}


@dataclass
class SizeColumn:
    source_col_idx: int
    source_header_text: str
    size_no: str


@dataclass
class DetectResult:
    sheet_name: str
    header_row: int
    color_col_idx: int
    color_header_text: str
    size_columns: List[SizeColumn]

    @property
    def color_display_text(self) -> str:
        if self.color_header_text == "款号":
            return "款号"
        return "款色"

    @property
    def color_info_text(self) -> str:
        if self.color_header_text == "款号":
            return "款号（按款色处理）"
        return self.color_header_text


class ExcelTransformError(Exception):
    pass


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ").replace("\r", " ")


def is_color_header(value) -> bool:
    return normalize_text(value) in COLOR_ALIASES


def extract_size_no(value) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, int):
        return str(value) if value > 0 else None

    if isinstance(value, float) and value.is_integer() and value > 0:
        return str(int(value))

    text = normalize_text(value).replace(" ", "")
    if not text:
        return None

    m = re.fullmatch(r"尺码(\d+)", text, flags=re.IGNORECASE)
    if m:
        return m.group(1)

    m = re.fullmatch(r"size(\d+)", text, flags=re.IGNORECASE)
    if m:
        return m.group(1)

    if re.fullmatch(r"\d+", text):
        return text

    return None


def copy_style(src_cell, dst_cell):
    if src_cell is None:
        return
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.number_format = src_cell.number_format


class WorkbookAnalyzer:
    @staticmethod
    def detect_header_row(ws) -> Optional[int]:
        best_row = None
        best_size_count = -1
        for row_idx in range(1, min(ws.max_row, HEADER_SCAN_MAX_ROW) + 1):
            color_found = False
            size_count = 0
            for col_idx in range(1, ws.max_column + 1):
                value = ws.cell(row_idx, col_idx).value
                if is_color_header(value):
                    color_found = True
                elif extract_size_no(value):
                    size_count += 1
            if color_found and size_count > best_size_count:
                best_row = row_idx
                best_size_count = size_count
        return best_row

    @staticmethod
    def detect_columns(ws, header_row: int) -> DetectResult:
        color_col_idx = None
        color_header_text = ""
        size_columns: List[SizeColumn] = []

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col_idx).value
            text = normalize_text(cell_value)

            if is_color_header(cell_value):
                color_col_idx = col_idx
                color_header_text = text
                continue

            size_no = extract_size_no(cell_value)
            if size_no is not None:
                size_columns.append(
                    SizeColumn(
                        source_col_idx=col_idx,
                        source_header_text=text or f"尺码{size_no}",
                        size_no=size_no,
                    )
                )

        if color_col_idx is None:
            raise ExcelTransformError("没有找到“款色”或“款号”列。")

        if not size_columns:
            raise ExcelTransformError("没有找到尺码列（支持：尺码1 / 1 / size1）。")

        size_columns.sort(key=lambda x: (int(x.size_no), x.source_col_idx))

        return DetectResult(
            sheet_name=ws.title,
            header_row=header_row,
            color_col_idx=color_col_idx,
            color_header_text=color_header_text,
            size_columns=size_columns,
        )


class PreviewBuilder:
    @staticmethod
    def get_group_page_info(size_columns: List[SizeColumn], page_index: int) -> Tuple[List[SizeColumn], int, int]:
        total_groups = len(size_columns)
        total_pages = max(1, math.ceil(total_groups / GROUPS_PER_PAGE))
        page_index = max(0, min(page_index, total_pages - 1))
        start = page_index * GROUPS_PER_PAGE
        end = min(start + GROUPS_PER_PAGE, total_groups)
        return size_columns[start:end], page_index, total_pages

    @staticmethod
    def build_source_preview(ws, detect: DetectResult, page_index: int):
        subset, page_index, total_pages = PreviewBuilder.get_group_page_info(detect.size_columns, page_index)
        columns = [detect.color_display_text] + [f"尺码{sc.size_no}" for sc in subset]
        rows = []

        for row_idx in range(detect.header_row + 1, ws.max_row + 1):
            color_text = normalize_text(ws.cell(row_idx, detect.color_col_idx).value)
            if not color_text:
                continue
            row = [color_text]
            for sc in subset:
                row.append(ws.cell(row_idx, sc.source_col_idx).value)
            rows.append(row)
            if len(rows) >= PREVIEW_DATA_ROWS:
                break

        return columns, rows, page_index, total_pages

    @staticmethod
    def build_result_preview(ws, detect: DetectResult, page_index: int):
        subset, page_index, total_pages = PreviewBuilder.get_group_page_info(detect.size_columns, page_index)
        # Row 1 / Row 2 exactly as final sheet
        row1 = [""]
        for sc in subset:
            row1.extend(["", f"尺码{sc.size_no}"])
        row2 = [detect.color_display_text]
        for sc in subset:
            row2.extend(["公式变成", sc.size_no])

        rows = [row1, row2]
        for row_idx in range(detect.header_row + 1, ws.max_row + 1):
            color_text = normalize_text(ws.cell(row_idx, detect.color_col_idx).value)
            if not color_text:
                continue
            data_row = [color_text]
            for sc in subset:
                data_row.append(f"{color_text}{sc.size_no}")
                data_row.append(ws.cell(row_idx, sc.source_col_idx).value)
            rows.append(data_row)
            if len(rows) >= PREVIEW_DATA_ROWS + 2:
                break

        return rows, subset, page_index, total_pages


class WorkbookTransformer:
    FORMULA_HEADER_FILL = PatternFill("solid", fgColor="FFFF00")

    @staticmethod
    def choose_result_sheet(wb, source_title: str):
        preferred = "变成"
        if preferred in wb.sheetnames and preferred != source_title:
            return wb[preferred], True

        result_name = preferred
        if result_name == source_title:
            result_name = f"{source_title}_处理结果"

        if result_name in wb.sheetnames:
            del wb[result_name]

        ws = wb.create_sheet(result_name)
        return ws, False

    @staticmethod
    def setup_generic_layout(result_ws, source_ws, detect: DetectResult):
        color_header_cell = source_ws.cell(detect.header_row, detect.color_col_idx)
        first_size_header = source_ws.cell(detect.header_row, detect.size_columns[0].source_col_idx)
        first_color_data = source_ws.cell(detect.header_row + 1, detect.color_col_idx) if detect.header_row + 1 <= source_ws.max_row else None
        first_size_data = source_ws.cell(detect.header_row + 1, detect.size_columns[0].source_col_idx) if detect.header_row + 1 <= source_ws.max_row else None

        result_ws.sheet_view.showGridLines = True
        result_ws.freeze_panes = "A3"

        result_ws.row_dimensions[1].height = 16.5
        result_ws.row_dimensions[2].height = 16.5

        # A column
        result_ws.column_dimensions["A"].width = max(source_ws.column_dimensions[get_column_letter(detect.color_col_idx)].width or 14, 12)
        # formula + size columns
        for idx, sc in enumerate(detect.size_columns):
            formula_col = 2 + idx * 2
            value_col = 3 + idx * 2
            formula_letter = get_column_letter(formula_col)
            value_letter = get_column_letter(value_col)

            source_width = source_ws.column_dimensions[get_column_letter(sc.source_col_idx)].width or 10
            result_ws.column_dimensions[formula_letter].width = max(result_ws.column_dimensions["A"].width, 12)
            result_ws.column_dimensions[value_letter].width = max(source_width, 9)

        # first row labels style
        for col_idx in range(1, 2 + len(detect.size_columns) * 2):
            cell = result_ws.cell(1, col_idx)
            cell.font = Font(name="微软雅黑", size=11, bold=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # second row headers
        copy_style(color_header_cell, result_ws.cell(2, 1))
        result_ws.cell(2, 1).value = detect.color_display_text

        for idx, sc in enumerate(detect.size_columns):
            formula_col = 2 + idx * 2
            value_col = 3 + idx * 2

            formula_header_cell = result_ws.cell(2, formula_col)
            copy_style(first_size_header, formula_header_cell)
            formula_header_cell.fill = copy(WorkbookTransformer.FORMULA_HEADER_FILL)
            formula_header_cell.value = "公式变成"

            value_header_cell = result_ws.cell(2, value_col)
            copy_style(first_size_header, value_header_cell)
            value_header_cell.value = int(sc.size_no) if sc.size_no.isdigit() else sc.size_no

        # default data row style
        for row_idx in range(3, source_ws.max_row - detect.header_row + 3):
            result_ws.row_dimensions[row_idx].height = 16.5

        if first_color_data:
            copy_style(first_color_data, result_ws.cell(3, 1))
        if first_size_data:
            copy_style(first_size_data, result_ws.cell(3, 3))

    @staticmethod
    def clear_result_area(result_ws, max_rows: int, max_cols: int):
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                result_ws.cell(row, col).value = None

    @staticmethod
    def fill_result_sheet(result_ws, source_ws, detect: DetectResult, preserve_existing_styles: bool):
        total_cols = 1 + len(detect.size_columns) * 2
        total_rows = max(3, source_ws.max_row - detect.header_row + 2)

        WorkbookTransformer.clear_result_area(result_ws, max(result_ws.max_row, total_rows + 5), max(total_cols, result_ws.max_column))

        if not preserve_existing_styles:
            WorkbookTransformer.setup_generic_layout(result_ws, source_ws, detect)

        # Row 1
        result_ws.cell(1, 1).value = None
        for idx, sc in enumerate(detect.size_columns):
            formula_col = 2 + idx * 2
            value_col = 3 + idx * 2
            result_ws.cell(1, formula_col).value = None
            c = result_ws.cell(1, value_col)
            c.value = f"尺码{sc.size_no}"
            if not preserve_existing_styles:
                c.font = Font(name="微软雅黑", size=11, bold=False)
                c.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2
        result_ws.cell(2, 1).value = detect.color_display_text
        for idx, sc in enumerate(detect.size_columns):
            formula_col = 2 + idx * 2
            value_col = 3 + idx * 2
            result_ws.cell(2, formula_col).value = "公式变成"
            result_ws.cell(2, value_col).value = int(sc.size_no) if sc.size_no.isdigit() else sc.size_no
            if preserve_existing_styles:
                # existing template sheet: keep style as is
                pass

        # Data rows
        result_row = 3
        for source_row in range(detect.header_row + 1, source_ws.max_row + 1):
            color_val = source_ws.cell(source_row, detect.color_col_idx).value
            color_text = normalize_text(color_val)
            if not color_text:
                continue

            color_dst = result_ws.cell(result_row, 1)
            color_dst.value = color_val
            if not preserve_existing_styles:
                copy_style(source_ws.cell(source_row, detect.color_col_idx), color_dst)

            for idx, sc in enumerate(detect.size_columns):
                formula_col = 2 + idx * 2
                value_col = 3 + idx * 2

                formula_dst = result_ws.cell(result_row, formula_col)
                value_dst = result_ws.cell(result_row, value_col)

                value_cell_src = source_ws.cell(source_row, sc.source_col_idx)
                formula_dst.value = f"=A{result_row}&${get_column_letter(value_col)}$2"
                value_dst.value = value_cell_src.value

                if not preserve_existing_styles:
                    copy_style(source_ws.cell(source_row, detect.color_col_idx), formula_dst)
                    copy_style(value_cell_src, value_dst)

            result_row += 1

        # If template existed and has more rows than current data, clear extras in used columns only
        for row in range(result_row, max(result_ws.max_row, result_row) + 1):
            if row >= result_row and row <= result_ws.max_row:
                for col in range(1, total_cols + 1):
                    result_ws.cell(row, col).value = None

    @staticmethod
    def process_file(input_path: str, output_path: str, sheet_name: str, header_row: int):
        wb = load_workbook(input_path)

        if sheet_name not in wb.sheetnames:
            raise ExcelTransformError(f"工作表不存在：{sheet_name}")

        source_ws = wb[sheet_name]
        detect = WorkbookAnalyzer.detect_columns(source_ws, header_row)

        result_ws, preserve = WorkbookTransformer.choose_result_sheet(wb, sheet_name)
        WorkbookTransformer.fill_result_sheet(result_ws, source_ws, detect, preserve_existing_styles=preserve)

        # Make result sheet active
        wb.active = wb.sheetnames.index(result_ws.title)

        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        wb.save(output_path)
        return detect, result_ws.title


class ResultPreviewCanvas(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.canvas = tk.Canvas(self, background="white", highlightthickness=0)
        self.v_scroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.h_scroll = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.v_scroll.grid(row=0, column=1, sticky="ns")
        self.h_scroll.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self._current_rows = []
        self._current_detect = None
        self._current_subset = []

    def render(self, rows: List[List[object]], detect: DetectResult, subset: List[SizeColumn]):
        self.canvas.delete("all")
        self._current_rows = rows
        self._current_detect = detect
        self._current_subset = subset

        if not rows:
            self.canvas.create_text(20, 20, text="暂无预览数据", anchor="nw", font=("微软雅黑", 10))
            self.canvas.configure(scrollregion=(0, 0, 400, 120))
            return

        col_widths = self._build_col_widths(subset)
        row_heights = [28, 28] + [28] * max(0, len(rows) - 2)

        x = 0
        y = 0
        total_w = sum(col_widths)
        total_h = sum(row_heights)

        for r_idx, row in enumerate(rows):
            x = 0
            for c_idx, width in enumerate(col_widths):
                value = row[c_idx] if c_idx < len(row) else ""
                fill = "white"
                text_color = "black"
                font = ("微软雅黑", 10)
                if r_idx == 1:
                    if c_idx == 0 or c_idx % 2 == 0:
                        fill = "#D9E2F3"
                        font = ("微软雅黑", 10, "bold")
                    else:
                        fill = "#FFFF00"
                        font = ("微软雅黑", 10, "bold")
                elif r_idx == 0:
                    fill = "white"
                    font = ("微软雅黑", 11)

                self.canvas.create_rectangle(x, y, x + width, y + row_heights[r_idx], fill=fill, outline="#C9CED6")
                self.canvas.create_text(
                    x + width / 2,
                    y + row_heights[r_idx] / 2,
                    text="" if value is None else str(value),
                    fill=text_color,
                    font=font,
                    anchor="center",
                )
                x += width
            y += row_heights[r_idx]

        self.canvas.configure(scrollregion=(0, 0, total_w, total_h))

    def _build_col_widths(self, subset: List[SizeColumn]) -> List[int]:
        widths = [120]
        for _ in subset:
            widths.extend([120, 70])
        return widths


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1720x980")
        self.minsize(1380, 820)

        self.input_path_var = tk.StringVar()
        self.output_path_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.header_row_var = tk.IntVar(value=1)
        self.confirm_detect_var = tk.BooleanVar(value=False)
        self.confirm_preview_var = tk.BooleanVar(value=False)

        self.wb_preview = None
        self.detect_result: Optional[DetectResult] = None
        self.preview_page_index = 0

        self._build_ui()

    def log(self, message: str):
        now = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{now}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _build_ui(self):
        container = ttk.Frame(self, padding=12)
        container.pack(fill="both", expand=True)

        # file settings
        file_frame = ttk.LabelFrame(container, text="文件设置", padding=10)
        file_frame.pack(fill="x")

        ttk.Label(file_frame, text="输入文件:").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Entry(file_frame, textvariable=self.input_path_var).grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Button(file_frame, text="选择文件", command=self.choose_input_file).grid(row=0, column=2, padx=6)
        ttk.Button(file_frame, text="加载工作簿", command=self.load_workbook_preview).grid(row=0, column=3, padx=6)

        ttk.Label(file_frame, text="输出文件:").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Entry(file_frame, textvariable=self.output_path_var).grid(row=1, column=1, sticky="ew", padx=8)
        ttk.Button(file_frame, text="选择保存位置", command=self.choose_output_file).grid(row=1, column=2, padx=6)
        file_frame.columnconfigure(1, weight=1)

        # settings
        settings_frame = ttk.LabelFrame(container, text="识别与处理设置", padding=10)
        settings_frame.pack(fill="x", pady=(10, 0))

        ttk.Label(settings_frame, text="工作表:").grid(row=0, column=0, sticky="w")
        self.sheet_combo = ttk.Combobox(settings_frame, textvariable=self.sheet_var, state="readonly", width=28)
        self.sheet_combo.grid(row=0, column=1, sticky="w", padx=(6, 18))

        ttk.Label(settings_frame, text="表头行:").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(settings_frame, from_=1, to=9999, textvariable=self.header_row_var, width=8).grid(
            row=0, column=3, sticky="w", padx=(6, 18)
        )

        ttk.Label(settings_frame, text="输出格式: 按原表风格复制，插入“公式变成”列").grid(row=0, column=4, sticky="w")

        btn_row = ttk.Frame(settings_frame)
        btn_row.grid(row=1, column=0, columnspan=5, sticky="w", pady=(10, 6))
        ttk.Button(btn_row, text="自动识别表头", command=self.auto_detect).pack(side="left", padx=(0, 6))
        ttk.Button(btn_row, text="生成预览", command=self.generate_preview).pack(side="left", padx=6)
        ttk.Button(btn_row, text="开始处理并保存", command=self.process_and_save).pack(side="left", padx=6)

        ttk.Checkbutton(settings_frame, text="我已核对识别结果", variable=self.confirm_detect_var).grid(
            row=2, column=0, columnspan=2, sticky="w"
        )
        ttk.Checkbutton(settings_frame, text="我已核对预览结果", variable=self.confirm_preview_var).grid(
            row=2, column=2, columnspan=2, sticky="w"
        )

        # detection / preview
        preview_frame = ttk.LabelFrame(container, text="识别结果与预览", padding=10)
        preview_frame.pack(fill="both", expand=True, pady=(10, 0))

        self.detect_text = tk.Text(preview_frame, height=5, wrap="word")
        self.detect_text.pack(fill="x")

        self.notebook = ttk.Notebook(preview_frame)
        self.notebook.pack(fill="both", expand=True, pady=(8, 0))

        # source preview tab
        source_tab = ttk.Frame(self.notebook)
        self.notebook.add(source_tab, text="源表内容预览")
        self.source_tree = ttk.Treeview(source_tab, show="headings")
        self.source_tree.pack(side="left", fill="both", expand=True)
        self.source_vscroll = ttk.Scrollbar(source_tab, orient="vertical", command=self.source_tree.yview)
        self.source_tree.configure(yscrollcommand=self.source_vscroll.set)
        self.source_vscroll.pack(side="right", fill="y")

        # result preview tab
        result_tab = ttk.Frame(self.notebook)
        self.notebook.add(result_tab, text="结果模板预览")

        top_bar = ttk.Frame(result_tab)
        top_bar.pack(fill="x", pady=(0, 6))
        self.page_label = ttk.Label(top_bar, text="预览页: 1/1")
        self.page_label.pack(side="left")
        self.prev_btn = ttk.Button(top_bar, text="上一页", command=self.prev_page)
        self.prev_btn.pack(side="right", padx=(6, 0))
        self.next_btn = ttk.Button(top_bar, text="下一页", command=self.next_page)
        self.next_btn.pack(side="right")

        self.result_canvas_preview = ResultPreviewCanvas(result_tab)
        self.result_canvas_preview.pack(fill="both", expand=True)

        # log
        log_frame = ttk.LabelFrame(container, text="日志", padding=10)
        log_frame.pack(fill="both", expand=False, pady=(10, 0))
        self.log_text = tk.Text(log_frame, height=8, state="disabled")
        self.log_text.pack(fill="both", expand=True)

        self.log("工具已启动。建议流程：选择文件 -> 加载工作簿 -> 自动识别 -> 生成预览 -> 勾选两项确认 -> 开始处理。")

    def choose_input_file(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        if not path:
            return
        self.input_path_var.set(path)
        if not self.output_path_var.get().strip():
            base, ext = os.path.splitext(path)
            self.output_path_var.set(f"{base}_处理结果{ext or '.xlsx'}")
        self.log(f"已选择输入文件：{path}")

    def choose_output_file(self):
        path = filedialog.asksaveasfilename(
            title="选择保存位置",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        self.output_path_var.set(path)
        self.log(f"已选择输出文件：{path}")

    def load_workbook_preview(self):
        input_path = self.input_path_var.get().strip()
        if not input_path:
            messagebox.showwarning(APP_TITLE, "请先选择输入文件。")
            return

        try:
            self.wb_preview = load_workbook(input_path, data_only=False)
            self.sheet_combo["values"] = self.wb_preview.sheetnames
            if self.wb_preview.sheetnames:
                self.sheet_var.set(self.wb_preview.sheetnames[0])
            self.detect_result = None
            self.preview_page_index = 0
            self.confirm_detect_var.set(False)
            self.confirm_preview_var.set(False)
            self.detect_text.delete("1.0", "end")
            self.clear_source_preview()
            self.result_canvas_preview.render([], None, [])
            self.log(f"工作簿加载成功，共 {len(self.wb_preview.sheetnames)} 个工作表：{', '.join(self.wb_preview.sheetnames)}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"加载工作簿失败：{e}")
            self.log(f"加载工作簿失败：{e}")

    def get_current_ws(self):
        if self.wb_preview is None:
            raise ExcelTransformError("请先加载工作簿。")
        sheet_name = self.sheet_var.get().strip()
        if not sheet_name:
            raise ExcelTransformError("请选择工作表。")
        if sheet_name not in self.wb_preview.sheetnames:
            raise ExcelTransformError(f"工作表不存在：{sheet_name}")
        return self.wb_preview[sheet_name]

    def auto_detect(self):
        try:
            ws = self.get_current_ws()
            header_row = WorkbookAnalyzer.detect_header_row(ws)
            if not header_row:
                raise ExcelTransformError("自动识别失败：前 20 行内没有同时找到“款色/款号”和尺码列。")
            self.header_row_var.set(header_row)
            self.detect_result = WorkbookAnalyzer.detect_columns(ws, header_row)
            self.show_detect_info(self.detect_result)
            self.confirm_detect_var.set(False)
            self.confirm_preview_var.set(False)
            self.log(f"自动识别完成：工作表={ws.title}，表头行={header_row}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            self.log(f"自动识别失败：{e}")

    def show_detect_info(self, detect: DetectResult):
        size_desc = "，".join([f"{get_column_letter(s.source_col_idx)}=尺码{s.size_no}" for s in detect.size_columns])
        text = (
            f"工作表：{detect.sheet_name}\n"
            f"表头行：{detect.header_row}\n"
            f"主列：{get_column_letter(detect.color_col_idx)}（{detect.color_info_text}）\n"
            f"识别到 {len(detect.size_columns)} 个尺码列：{size_desc}\n"
            f"说明：如果识别到“款号”，会按“款色”处理；结果页第1行在数字列正上方显示“尺码X”。"
        )
        self.detect_text.delete("1.0", "end")
        self.detect_text.insert("1.0", text)

    def clear_source_preview(self):
        self.source_tree.delete(*self.source_tree.get_children())
        self.source_tree["columns"] = ()

    def render_source_preview(self, ws, detect):
        columns, rows, page_index, total_pages = PreviewBuilder.build_source_preview(ws, detect, self.preview_page_index)
        self.clear_source_preview()
        self.source_tree["columns"] = [f"c{i}" for i in range(len(columns))]
        for i, title in enumerate(columns):
            key = f"c{i}"
            self.source_tree.heading(key, text=title)
            self.source_tree.column(key, width=120 if i == 0 else 90, anchor="center")
        for row in rows:
            self.source_tree.insert("", "end", values=row)

    def render_result_preview(self, ws, detect):
        rows, subset, page_index, total_pages = PreviewBuilder.build_result_preview(ws, detect, self.preview_page_index)
        self.page_label.configure(text=f"预览页: {page_index + 1}/{total_pages}")
        self.prev_btn.configure(state="normal" if page_index > 0 else "disabled")
        self.next_btn.configure(state="normal" if page_index < total_pages - 1 else "disabled")
        self.result_canvas_preview.render(rows, detect, subset)

    def generate_preview(self):
        try:
            ws = self.get_current_ws()
            header_row = int(self.header_row_var.get())
            self.detect_result = WorkbookAnalyzer.detect_columns(ws, header_row)
            self.show_detect_info(self.detect_result)
            self.preview_page_index = 0
            self.render_source_preview(ws, self.detect_result)
            self.render_result_preview(ws, self.detect_result)
            self.confirm_preview_var.set(False)
            self.log("预览已生成。结果模板预览区直接按最终 Excel 的双层表头结构展示。")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            self.log(f"生成预览失败：{e}")

    def prev_page(self):
        if self.detect_result is None:
            return
        self.preview_page_index = max(0, self.preview_page_index - 1)
        ws = self.get_current_ws()
        self.render_source_preview(ws, self.detect_result)
        self.render_result_preview(ws, self.detect_result)

    def next_page(self):
        if self.detect_result is None:
            return
        self.preview_page_index += 1
        ws = self.get_current_ws()
        self.render_source_preview(ws, self.detect_result)
        self.render_result_preview(ws, self.detect_result)

    def process_and_save(self):
        input_path = self.input_path_var.get().strip()
        output_path = self.output_path_var.get().strip()

        if not input_path:
            messagebox.showwarning(APP_TITLE, "请先选择输入文件。")
            return
        if not output_path:
            messagebox.showwarning(APP_TITLE, "请选择输出保存位置。")
            return
        if self.detect_result is None:
            messagebox.showwarning(APP_TITLE, "请先自动识别或生成预览。")
            return
        if not self.confirm_detect_var.get():
            messagebox.showwarning(APP_TITLE, "请先勾选“我已核对识别结果”。")
            return
        if not self.confirm_preview_var.get():
            messagebox.showwarning(APP_TITLE, "请先勾选“我已核对预览结果”。")
            return

        summary = (
            f"即将处理：\n"
            f"- 文件：{os.path.basename(input_path)}\n"
            f"- 工作表：{self.detect_result.sheet_name}\n"
            f"- 表头行：{self.detect_result.header_row}\n"
            f"- 主列：{get_column_letter(self.detect_result.color_col_idx)}（{self.detect_result.color_info_text}）\n"
            f"- 尺码数量：{len(self.detect_result.size_columns)}\n"
            f"- 保存到：{output_path}\n\n"
            f"确认继续吗？"
        )
        if not messagebox.askyesno(APP_TITLE, summary):
            self.log("用户取消了处理。")
            return

        try:
            detect, result_sheet_name = WorkbookTransformer.process_file(
                input_path=input_path,
                output_path=output_path,
                sheet_name=self.detect_result.sheet_name,
                header_row=self.detect_result.header_row,
            )
            messagebox.showinfo(APP_TITLE, f"处理完成。\n结果工作表：{result_sheet_name}\n输出文件：{output_path}")
            self.log(f"已生成结果工作表：{result_sheet_name}")
            self.log(f"文件已保存到：{output_path}")
        except Exception as e:
            self.log(f"处理失败：{e}")
            self.log(traceback.format_exc())
            messagebox.showerror(APP_TITLE, f"处理失败：{e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
